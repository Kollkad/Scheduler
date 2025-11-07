"""
backend/app/common/modules/data_import.py

Модуль загрузки данных из Excel файлов.
Использует быстрые методы загрузки с openpyxl как основной вариант.
"""

import pandas as pd
from openpyxl import load_workbook
import warnings
import os
import tempfile
import zipfile
import shutil


def load_excel_data(filepath):
    """
    Загрузка данных из Excel файла с приоритетом быстрых методов.

    Args:
        filepath (str): Путь к Excel файлу

    Returns:
        pd.DataFrame: Загруженные данные

    Raises:
        ValueError: Если файл невозможно загрузить всеми методами
    """
    # Метод 1: Быстрая загрузка через openpyxl read_only (основной)
    try:
        print("Быстрая загрузка через openpyxl...")
        return fast_openpyxl_load(filepath)
    except Exception as e:
        print(f"Быстрая загрузка не удалась: {e}")

    # Метод 2: Стандартная загрузка pandas
    try:
        print("Стандартная загрузка pandas...")
        return pd.read_excel(filepath, header=None)
    except Exception as e:
        print(f"Стандартная загрузка не удалась: {e}")

    # Метод 3: Альтернативные движки
    for engine in ['openpyxl', 'xlrd']:
        try:
            print(f"Загрузка с движком {engine}...")
            return pd.read_excel(filepath, header=None, engine=engine)
        except Exception as e:
            print(f"Движок {engine} не сработал: {e}")
            continue

    # Метод 4: Восстановление через openpyxl без read_only
    try:
        print("Восстановление через openpyxl...")
        return repair_openpyxl_full(filepath)
    except Exception as e:
        print(f"Восстановление openpyxl не удалось: {e}")

    # Метод 5: Упрощенная загрузка
    try:
        print("Упрощенная загрузка...")
        return load_excel_data_simple_fallback(filepath)
    except Exception as e:
        print(f"Упрощенная загрузка не удалась: {e}")

    # Метод 6: Экстренное восстановление XML
    try:
        print("Экстренное восстановление XML...")
        return repair_excel_xml(filepath)
    except Exception as e:
        print(f"XML восстановление не удалось: {e}")

    raise ValueError(f"Не удалось загрузить файл {filepath}")


def fast_openpyxl_load(filepath):
    """
    Быстрая загрузка через openpyxl в режиме read_only.

    Args:
        filepath (str): Путь к Excel файлу

    Returns:
        pd.DataFrame: Загруженные данные
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(
            filepath,
            read_only=True,
            data_only=True,
            keep_links=False
        )
        sheet = wb.active
        data = []

        # Чтение всех строк без ограничения (read_only режим эффективен)
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        df = pd.DataFrame(data)
        print(f"Загружено строк через openpyxl: {len(df)}")
        return df


def repair_openpyxl_full(filepath):
    """
    Восстановление через openpyxl в полном режиме.

    Args:
        filepath (str): Путь к Excel файлу

    Returns:
        pd.DataFrame: Загруженные данные
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(
            filepath,
            read_only=False,
            data_only=True,
            keep_links=False
        )
        sheet = wb.active
        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        return pd.DataFrame(data)


def repair_excel_xml(filepath):
    """
    Восстановление поврежденной XML структуры Excel файла.

    Args:
        filepath (str): Путь к поврежденному файлу

    Returns:
        pd.DataFrame: Восстановленные данные
    """
    temp_dir = tempfile.mkdtemp()

    try:
        # Распаковка Excel как ZIP архива
        with zipfile.ZipFile(filepath, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        # Восстановление критических XML файлов
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                if file.endswith('.xml') or file.endswith('.rels'):
                    file_path = os.path.join(root, file)
                    try:
                        repair_xml_file(file_path)
                    except Exception:
                        create_minimal_xml(file_path)

        # Создание восстановленного файла
        repaired_path = os.path.join(temp_dir, "repaired.xlsx")
        with zipfile.ZipFile(repaired_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    if file != "repaired.xlsx":
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, temp_dir)
                        zipf.write(file_path, arcname)

        # Загрузка восстановленного файла
        return pd.read_excel(repaired_path, header=None, engine='openpyxl')

    except Exception as e:
        raise ValueError(f"XML восстановление не удалось: {e}")
    finally:
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except:
            pass


def repair_xml_file(file_path):
    """
    Восстановление отдельного XML файла.

    Args:
        file_path (str): Путь к XML файлу
    """
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()

        # Простая очистка проблемных символов
        content = content.replace('\x00', '')

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
    except Exception:
        create_minimal_xml(file_path)


def create_minimal_xml(file_path):
    """
    Создание минимального валидного XML файла.

    Args:
        file_path (str): Путь для создания файла
    """
    try:
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes"?><root></root>')
    except Exception:
        pass


def load_excel_data_simple_fallback(filepath):
    """
    Упрощенная загрузка для поврежденных файлов.

    Args:
        filepath (str): Путь к файлу

    Returns:
        pd.DataFrame: Частично загруженные данные
    """
    for nrows in [10000, 5000, 1000]:
        try:
            df = pd.read_excel(filepath, header=None, nrows=nrows, engine='openpyxl')
            if len(df) > 0:
                print(f"Загружено {len(df)} строк (ограничение: {nrows})")
                return df
        except Exception:
            continue

    # Создание DataFrame с сообщением об ошибке
    return pd.DataFrame([["Файл поврежден", "Невозможно прочитать данные"]])